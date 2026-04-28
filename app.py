#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import streamlit as st
import pandas as pd
import numpy as np
import re
from io import BytesIO
from datetime import datetime, date, time, timedelta
from openpyxl import load_workbook

# --------------------------------------------------
# Streamlit config
# --------------------------------------------------

st.set_page_config(page_title="AuDi Flavor Predictor", layout="wide")

st.title("AuDi Flavor Probability Predictor")
st.markdown(
    "Upload Audi parametric data and AuDi sensory output file. "
    "The app processes, merges, scores batches, and supports what-if sliders."
)

# --------------------------------------------------
# Model config
# Coefficients are for SCALED variables
# --------------------------------------------------

MODELS = {
    "Fruity/Floral": {
        "intercept": 2.0550,
        "threshold": 0.863994,
        "variables": {
            "Acetaldehyde (0-6.5)(gm/100LAA)": {
                "coef": -1.401050,
                "mean": 2.703928571428572,
                "std": 2.824616079204032,
            },
            "Chloride (Max)": {
                "coef": -0.587540,
                "mean": 25.758928571428573,
                "std": 10.596702263064714,
            },
            "Acetal (0-6.0)(gm/100LAA)": {
                "coef": 0.423513,
                "mean": 1.4782142857142857,
                "std": 1.121263010981776,
            },
        },
    },

    "Cereal/Grainy": {
        "intercept": 0.6548,
        "threshold": 0.141238,
        "variables": {
            "Yeast Storage Temperature, deg C": {
                "coef": 0.385188,
                "mean": 21.910714285714285,
                "std": 0.7855113374258131,
            },
            "Wash condenser temperature(Max) - Mean_Wash_Still_A": {
                "coef": 1.447936,
                "mean": 38.15000000000001,
                "std": 3.2777115274445454,
            },
            "Sparging water temperature": {
                "coef": 0.688265,
                "mean": 78.19821428571429,
                "std": 1.038627368801963,
            },
        },
    },

    "Starchy": {
        "intercept": 0.0000,
        "threshold": 0.261504,
        "variables": {
            "Final Wash temperature_Fermentated Wash": {
                "coef": 0.345272,
                "mean": 32.348214285714285,
                "std": 0.9974057548439944,
            },
            "Fermented wash residual sugar_Fermentated Wash": {
                "coef": 0.159157,
                "mean": 0.4192857142857143,
                "std": 0.048948119140033754,
            },
        },
    },

    "Fermented": {
        "intercept": 0.3048,
        "threshold": 0.116056,
        "variables": {
            "Fermentation time": {
                "coef": -0.588327,
                "mean": 12.240625,
                "std": 6.21380462861117,
            },
            "Alkalinity": {
                "coef": 0.534245,
                "mean": 71.07142857142857,
                "std": 17.619851976491443,
            },
            "Yeast Storage Temperature, deg C": {
                "coef": -0.352033,
                "mean": 21.910714285714285,
                "std": 0.7855113374258131,
            },
        },
    },

    "Cooked": {
        "intercept": -0.0011,
        "threshold": 0.26690,
        "variables": {
            "Recovery of FMS_Spirit_B": {
                "coef": 0.297182,
                "mean": 2063.973214285714,
                "std": 37.69265034019719,
            },
            "Spirit distillation time_Spirit_A": {
                "coef": 0.511045,
                "mean": 9.93675595238095,
                "std": 0.5984691880223978,
            },
            "Malt foreign matter": {
                "coef": 0.339665,
                "mean": 0.2774107142857143,
                "std": 0.06459690306202055,
            },
        },
    },

    "Acidic/Solvent": {
        "intercept": 0.0000,
        "threshold": 0.435695,
        "variables": {
            "Ethyl Acetate (25-60)(gm/100LAA)": {
                "coef": -0.132881,
                "mean": 41.28142857142858,
                "std": 7.304673633212856,
            },
        },
    },
}

# --------------------------------------------------
# Parametric preprocessing helpers
# --------------------------------------------------

TABS = [
    "mashing",
    "Fermentation",
    "Wash Still A",
    "wash Still B",
    "Spirit still A",
    "Spirit Still B",
]

TIME_FORMAT_TOKENS = ("h:", "hh:", "[h]", ":mm", ":ss", "am/pm")


def clean(v):
    if v is None:
        return None
    if isinstance(v, str):
        t = v.replace("\xa0", " ").strip()
        return None if t == "" else t
    return v


def parameter_name(ws, row):
    if ws.title.lower() == "mashing":
        return clean(ws.cell(row=row, column=1).value), None

    ptype = clean(ws.cell(row=row, column=1).value)
    pname = clean(ws.cell(row=row, column=2).value)
    qualifier = None

    if isinstance(ptype, str) and ":" in ptype:
        qualifier = ptype.split(":")[-1].strip().replace(" ", "_").replace("-", "_")
        qualifier = "_".join(x for x in qualifier.split("_") if x)

    return pname, qualifier


def is_time_format(fmt):
    f = str(fmt or "").lower()
    return any(t in f for t in TIME_FORMAT_TOKENS)


def is_date_format(fmt):
    f = str(fmt or "").lower()
    return not is_time_format(f) and (
        "yy" in f or "yyyy" in f or "dd" in f or "mmm" in f
    )


def semantic_hint(name):
    n = str(name or "").lower().strip()

    if "date of" in n or n.startswith("date ") or (
        "date" in n and any(x in n for x in ["(a)", "(b)", "(c)"])
    ):
        return "date"

    if "shelf" in n and "life" in n:
        return "ratio"

    if "%" in n or "percentage" in n or "ratio" in n or "available" in n:
        return "ratio"

    if (
        "distillation time" in n
        or "time taken" in n
        or "duration" in n
        or (" time" in n and "date" not in n)
    ):
        return "hours"

    return "general"


def time_seconds_from_value(v):
    v = clean(v)

    if v is None:
        return None

    if isinstance(v, timedelta):
        return v.total_seconds()

    if isinstance(v, datetime):
        return v.hour * 3600 + v.minute * 60 + v.second

    if isinstance(v, time):
        return v.hour * 3600 + v.minute * 60 + v.second

    if isinstance(v, str):
        s = v.strip().lower()

        m = re.match(
            r"^(-?\d+(?:\.\d+)?)\s*(?:hr|hrs|hour|hours|h)\s*(?:(\d+(?:\.\d+)?)\s*(?:min|mins|minute|minutes|m))?$",
            s,
        )
        if m:
            return float(m.group(1)) * 3600 + float(m.group(2) or 0) * 60

        m = re.match(r"^(\d{1,3}):(\d{1,2})(?::(\d{1,2}(?:\.\d+)?))?$", s)
        if m:
            return int(m.group(1)) * 3600 + int(m.group(2)) * 60 + float(m.group(3) or 0)

    return None


def parse_date(v):
    v = clean(v)

    if v is None:
        return None

    if isinstance(v, datetime):
        return v

    if isinstance(v, date):
        return datetime(v.year, v.month, v.day)

    if isinstance(v, (int, float)) and 20000 <= float(v) <= 60000:
        return datetime(1899, 12, 30) + timedelta(days=float(v))

    if isinstance(v, str):
        for fmt in (
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d",
            "%d-%b-%y",
            "%d-%b-%Y",
            "%m-%d-%y",
            "%m/%d/%Y",
            "%d/%m/%Y",
        ):
            try:
                return datetime.strptime(v.strip(), fmt)
            except ValueError:
                pass

    return v


def normalize_ratio(v, source_format=None):
    v = clean(v)

    if v is None:
        return None

    seconds = time_seconds_from_value(v)

    if seconds is not None:
        return seconds / 86400

    if isinstance(v, str):
        try:
            v = float(v.replace("%", "").strip())
        except ValueError:
            return v

    if isinstance(v, (int, float)):
        x = float(v)
        if abs(x) > 1 and abs(x) <= 100:
            return x / 100
        return x

    return v


def normalize_hours(v, source_format=None):
    v = clean(v)

    if v is None:
        return None

    seconds = time_seconds_from_value(v)

    if seconds is not None:
        return seconds / 3600

    if isinstance(v, str):
        try:
            v = float(v.strip())
        except ValueError:
            return v

    if isinstance(v, (int, float)):
        x = float(v)
        if is_time_format(source_format) and 0 <= x < 1:
            return x * 24
        return x

    return v


def normalize_value(v, kind, source_format=None):
    if kind == "date":
        return parse_date(v)

    if kind == "ratio":
        return normalize_ratio(v, source_format)

    if kind == "hours":
        return normalize_hours(v, source_format)

    return clean(v)


def infer_kind(name, values, formats):
    hint = semantic_hint(name)

    if hint != "general":
        return hint

    nonblank = [v for v in values if clean(v) is not None]

    if not nonblank:
        return "general"

    time_like = sum(
        1
        for v, f in zip(values, formats)
        if clean(v) is not None and (is_time_format(f) or time_seconds_from_value(v) is not None)
    )

    date_like = sum(
        1
        for v, f in zip(values, formats)
        if clean(v) is not None and is_date_format(f)
    )

    if date_like >= max(2, int(len(nonblank) * 0.6)):
        return "date"

    if time_like >= max(2, int(len(nonblank) * 0.6)):
        return "hours"

    return "general"


def hhmm_number_to_decimal_hours(num):
    hrs = int(num)
    mins = round((float(num) - hrs) * 100)

    if 0 <= mins < 60:
        return hrs + mins / 60

    return num


def excel_1900_datetime_to_raw_number(x):
    day_part = x.day
    time_part = x.hour / 24 + x.minute / 1440 + x.second / 86400
    return day_part + time_part


def convert_process_time_to_hours(x):
    if pd.isna(x):
        return x

    if isinstance(x, pd.Timestamp):
        if x.year == 1900:
            raw_num = excel_1900_datetime_to_raw_number(x)
            return hhmm_number_to_decimal_hours(raw_num)
        return x

    if isinstance(x, datetime):
        if x.year == 1900:
            raw_num = excel_1900_datetime_to_raw_number(x)
            return hhmm_number_to_decimal_hours(raw_num)
        return x

    if isinstance(x, time):
        raw_num = x.hour + x.minute / 100 + x.second / 10000
        return hhmm_number_to_decimal_hours(raw_num)

    if isinstance(x, pd.Timedelta):
        return x.total_seconds() / 3600

    if isinstance(x, (int, float, np.integer, np.floating)):
        return hhmm_number_to_decimal_hours(x)

    s = str(x).strip()

    dt = pd.to_datetime(s, errors="coerce")

    if pd.notna(dt) and dt.year == 1900:
        raw_num = excel_1900_datetime_to_raw_number(dt)
        return hhmm_number_to_decimal_hours(raw_num)

    if ":" in s and "-" not in s:
        parts = s.split(":")
        try:
            hrs = int(parts[0])
            mins = int(parts[1])
            secs = int(float(parts[2])) if len(parts) > 2 else 0
            return hrs + mins / 60 + secs / 3600
        except Exception:
            return x

    if "." in s and s.replace(".", "", 1).isdigit():
        return hhmm_number_to_decimal_hours(float(s))

    return x


def looks_like_time_series(series):
    sample = series.dropna().astype(str).head(10)

    if len(sample) == 0:
        return False

    time_like = 0

    for val in sample:
        if ":" in val:
            time_like += 1
        elif "." in val:
            try:
                num = float(val)
                mins = (num - int(num)) * 100
                if 0 <= mins < 60:
                    time_like += 1
            except Exception:
                pass

    return time_like >= len(sample) * 0.5


def to_excel_date(x):
    if pd.isna(x):
        return pd.NaT

    if isinstance(x, pd.Timestamp):
        return x.normalize()

    if isinstance(x, (int, float, np.integer, np.floating)):
        return pd.to_datetime("1899-12-30") + pd.to_timedelta(float(x), unit="D")

    return pd.to_datetime(x, errors="coerce").normalize()


def find_matching_date_col(df, product, marker, suffix):
    product = product.lower()
    marker = marker.lower()

    candidates = []

    for col in df.columns:
        c = col.lower()

        if marker in c and product in c:
            if suffix and col.endswith(suffix):
                return col

            candidates.append(col)

    return candidates[0] if candidates else None


def fix_shelf_life_from_dates(df):
    df = df.copy()
    fixed_cols = []

    for col in df.columns:
        col_l = col.lower()

        if (
            "shelf-life at the time of use" not in col_l
            and "shelf life at the time of use" not in col_l
        ):
            continue

        product_match = re.search(r"\((.*?)\)", col)

        if not product_match:
            continue

        product = product_match.group(1)

        suffix = ""

        if "_" in col:
            suffix = "_" + col.split("_")[-1]

        mfg_col = find_matching_date_col(df, product, "date of mfg", suffix)
        exp_col = find_matching_date_col(df, product, "date of expiry", suffix)
        use_col = find_matching_date_col(df, product, "date of use", suffix)

        if not all([mfg_col, exp_col, use_col]):
            continue

        mfg = df[mfg_col].apply(to_excel_date)
        exp = df[exp_col].apply(to_excel_date)
        use = df[use_col].apply(to_excel_date)

        denom = (exp - mfg).dt.days
        numer = (exp - use).dt.days

        recalculated = np.where(
            (denom.notna()) & (denom != 0) & (numer.notna()),
            (numer / denom) * 100,
            df[col],
        )

        df[col] = recalculated

        fixed_cols.append(
            {
                "shelf_life_col": col,
                "mfg_col": mfg_col,
                "expiry_col": exp_col,
                "use_col": use_col,
            }
        )

    return df, pd.DataFrame(fixed_cols)


@st.cache_data(show_spinner=False)
def process_parametric_file(file_bytes):
    wb_f = load_workbook(BytesIO(file_bytes), data_only=False)
    wb_v = load_workbook(BytesIO(file_bytes), data_only=True)

    row_kind = {}

    available_tabs = wb_f.sheetnames
    missing_tabs = [tab for tab in TABS if tab not in available_tabs]

    if missing_tabs:
        raise ValueError(f"Missing required tabs in parametric file: {missing_tabs}")

    for tab in TABS:
        ws_f = wb_f[tab]
        ws_v = wb_v[tab]

        for row in range(3, ws_f.max_row + 1):
            pname, _ = parameter_name(ws_f, row)

            if pname is None:
                continue

            vals = []
            fmts = []

            for col in range(6, ws_f.max_column + 1):
                if clean(ws_f.cell(row=1, column=col).value) is None:
                    continue

                val = ws_v.cell(row=row, column=col).value

                if val is None:
                    val = ws_f.cell(row=row, column=col).value

                vals.append(val)
                fmts.append(ws_f.cell(row=row, column=col).number_format)

            row_kind[(tab, row)] = infer_kind(pname, vals, fmts)

    batches = []
    seen = set()

    for tab in TABS:
        ws = wb_f[tab]

        for col in range(6, ws.max_column + 1):
            batch = clean(ws.cell(row=1, column=col).value)

            if batch is not None and batch not in seen:
                seen.add(batch)
                batches.append(batch)

    combined_rows = {b: {"Sample": b} for b in batches}
    columns = ["Sample"]
    seen_cols = {"Sample"}

    for tab in TABS:
        ws_f = wb_f[tab]
        ws_v = wb_v[tab]

        tab_batches = {
            clean(ws_f.cell(row=1, column=col).value): col
            for col in range(6, ws_f.max_column + 1)
            if clean(ws_f.cell(row=1, column=col).value) is not None
        }

        for row in range(3, ws_f.max_row + 1):
            pname, qualifier = parameter_name(ws_f, row)

            if pname is None:
                continue

            base = pname

            if qualifier and qualifier.lower() not in pname.lower() and "_" not in pname:
                base = f"{pname}_{qualifier}"

            col_name = base

            if col_name in seen_cols:
                safe_tab = tab.replace(" ", "_")
                col_name = f"{base}_{safe_tab}"
                i = 2

                while col_name in seen_cols:
                    col_name = f"{base}_{safe_tab}_{i}"
                    i += 1

            seen_cols.add(col_name)
            columns.append(col_name)

            kind = row_kind.get((tab, row), "general")

            for b in batches:
                source_col = tab_batches.get(b)
                val = None

                if source_col:
                    fmt = ws_f.cell(row=row, column=source_col).number_format
                    val = ws_v.cell(row=row, column=source_col).value

                    if val is None:
                        val = ws_f.cell(row=row, column=source_col).value

                    val = normalize_value(val, kind, fmt)

                combined_rows[b][col_name] = val

    df_final = pd.DataFrame(
        [[combined_rows[b].get(c) for c in columns] for b in batches],
        columns=columns,
    )

    df_final, shelf_life_audit = fix_shelf_life_from_dates(df_final)

    process_time_cols = [
        col
        for col in df_final.columns
        if "time" in col.lower()
        and "date" not in col.lower()
        and looks_like_time_series(df_final[col])
    ]

    for col in process_time_cols:
        df_final[col] = df_final[col].apply(convert_process_time_to_hours)

    if "Recovery of FMS_Spirit_B" in df_final.columns:
        df_final.loc[
            df_final["Recovery of FMS_Spirit_B"] == 20615,
            "Recovery of FMS_Spirit_B",
        ] = 2061.5

    if "Wash distillation time_Wash_B" in df_final.columns:
        df_final.loc[
            df_final["Wash distillation time_Wash_B"] == "9.:50",
            "Wash distillation time_Wash_B",
        ] = 9.83

        df_final["Wash distillation time_Wash_B"] = pd.to_numeric(
            df_final["Wash distillation time_Wash_B"],
            errors="coerce",
        )

    df_final["Sample"] = df_final["Sample"].astype(str)

    return df_final, shelf_life_audit


@st.cache_data(show_spinner=False)
def process_sensory_file(file_bytes):
    sensory_df = pd.read_excel(
        BytesIO(file_bytes),
        sheet_name="TC Consensus Score-FMS",
        header=1,
    )

    if "Blinding_Code" not in sensory_df.columns:
        raise ValueError("Sensory file must contain 'Blinding_Code' column.")

    sensory_df["Blinding_Code_Base"] = (
        sensory_df["Blinding_Code"].astype(str).str.split("-").str[0].astype(str)
    )

    return sensory_df


def merge_parametric_and_sensory(parametric_df, sensory_df):
    if "Sample" not in parametric_df.columns:
        raise ValueError("Processed parametric data must contain 'Sample' column.")

    if "Blinding_Code_Base" not in sensory_df.columns:
        raise ValueError("Processed sensory data must contain 'Blinding_Code_Base' column.")

    merged = pd.merge(
        parametric_df,
        sensory_df,
        left_on="Sample",
        right_on="Blinding_Code_Base",
        how="inner",
    )

    return merged


# --------------------------------------------------
# Prediction helpers
# --------------------------------------------------

def sigmoid(z):
    z = np.clip(z, -500, 500)
    return 1 / (1 + np.exp(-z))


def scaled_value(raw_value, mean, std):
    if pd.isna(raw_value):
        raw_value = mean

    return (raw_value - mean) / std


def predict_from_values(flavor, values):
    model = MODELS[flavor]
    z = model["intercept"]

    for variable, params in model["variables"].items():
        raw_value = values[variable]
        scaled = scaled_value(raw_value, params["mean"], params["std"])
        z += params["coef"] * scaled

    probability = sigmoid(z)
    prediction = int(probability >= model["threshold"])

    return probability, prediction, z


def score_dataframe(df, flavor):
    model = MODELS[flavor]
    scored = df.copy()

    z = np.repeat(model["intercept"], len(scored)).astype(float)

    for variable, params in model["variables"].items():
        raw = pd.to_numeric(scored[variable], errors="coerce").fillna(params["mean"])
        scaled = (raw - params["mean"]) / params["std"]
        z += params["coef"] * scaled

    prob_col = f"{flavor}_Probability"
    pred_col = f"{flavor}_Prediction"

    scored[prob_col] = sigmoid(z)
    scored[pred_col] = (scored[prob_col] >= model["threshold"]).astype(int)

    return scored


# --------------------------------------------------
# Upload UI
# --------------------------------------------------

col1, col2 = st.columns(2)

with col1:
    parametric_file = st.file_uploader(
        "Upload Audi_parametric_data.xlsx",
        type=["xlsx"],
    )

with col2:
    sensory_file = st.file_uploader(
        "Upload AuDi_Sensory Output File.xlsx",
        type=["xlsx"],
    )


if parametric_file and sensory_file:
    try:
        with st.spinner("Processing parametric file..."):
            parametric_bytes = parametric_file.getvalue()
            parametric_df, shelf_life_audit = process_parametric_file(parametric_bytes)

        with st.spinner("Processing sensory file..."):
            sensory_bytes = sensory_file.getvalue()
            sensory_df = process_sensory_file(sensory_bytes)

        with st.spinner("Merging files..."):
            df = merge_parametric_and_sensory(parametric_df, sensory_df)

        st.success(f"Processing complete. Merged rows: {len(df)}")

        with st.expander("Preview processed parametric data"):
            st.dataframe(parametric_df.head())

        with st.expander("Preview sensory data"):
            st.dataframe(sensory_df.head())

        with st.expander("Preview merged modeling data"):
            st.dataframe(df.head())

        if not shelf_life_audit.empty:
            with st.expander("Shelf-life recalculation audit"):
                st.dataframe(shelf_life_audit)

        tabs = st.tabs(list(MODELS.keys()))

        for tab, flavor in zip(tabs, MODELS.keys()):
            with tab:
                model = MODELS[flavor]
                required_cols = list(model["variables"].keys())

                st.subheader(f"{flavor}")

                missing_cols = [c for c in required_cols if c not in df.columns]

                if missing_cols:
                    st.error("Missing required columns for this flavor:")
                    st.write(missing_cols)
                    continue

                scored_df = score_dataframe(df, flavor)

                st.markdown("### Batch Scoring")

                display_cols = []

                for col in ["Sample", "Blinding_Code", "Blinding_Code_Base"]:
                    if col in scored_df.columns:
                        display_cols.append(col)

                display_cols += required_cols
                display_cols += [
                    f"{flavor}_Probability",
                    f"{flavor}_Prediction",
                ]

                st.dataframe(scored_df[display_cols])

                csv = scored_df.to_csv(index=False).encode("utf-8")

                st.download_button(
                    label=f"Download {flavor} scored data",
                    data=csv,
                    file_name=f"{flavor.replace('/', '_').replace('&', 'and')}_scored.csv",
                    mime="text/csv",
                    key=f"{flavor}_download",
                )

                st.divider()
                st.markdown("### What-if Sliders")

                selected_index = st.selectbox(
                    "Select batch row for slider defaults",
                    scored_df.index,
                    key=f"{flavor}_row",
                )

                selected_row = scored_df.loc[selected_index]
                input_values = {}

                for variable in required_cols:
                    col_data = pd.to_numeric(df[variable], errors="coerce")

                    min_val = float(col_data.min())
                    max_val = float(col_data.max())

                    default_val = selected_row[variable]

                    if pd.isna(default_val):
                        default_val = model["variables"][variable]["mean"]

                    default_val = float(np.clip(default_val, min_val, max_val))

                    if min_val == max_val:
                        st.info(f"{variable}: constant value {min_val}")
                        input_values[variable] = min_val
                    else:
                        step = (max_val - min_val) / 100

                        input_values[variable] = st.slider(
                            variable,
                            min_value=min_val,
                            max_value=max_val,
                            value=default_val,
                            step=float(step),
                            key=f"{flavor}_{variable}",
                        )

                probability, prediction, z = predict_from_values(flavor, input_values)

                c1, c2, c3 = st.columns(3)

                with c1:
                    st.metric("Probability", f"{probability:.2%}")

                with c2:
                    st.metric("Threshold", f"{model['threshold']:.3f}")

                with c3:
                    if prediction == 1:
                        st.success("Outcome: 1")
                    else:
                        st.error("Outcome: 0")

                st.progress(float(probability))

                with st.expander("Calculation details"):
                    st.write(f"Logit z: {z:.4f}")

                    raw_inputs = pd.DataFrame(
                        {
                            "Variable": list(input_values.keys()),
                            "Raw Value": list(input_values.values()),
                        }
                    )

                    scaled_inputs = []

                    for variable in required_cols:
                        params = model["variables"][variable]
                        raw = input_values[variable]
                        scaled = scaled_value(raw, params["mean"], params["std"])

                        scaled_inputs.append(
                            {
                                "Variable": variable,
                                "Raw Value": raw,
                                "Training Mean": params["mean"],
                                "Training Std Dev": params["std"],
                                "Scaled Value Used": scaled,
                                "Coefficient": params["coef"],
                            }
                        )

                    st.dataframe(pd.DataFrame(scaled_inputs))

    except Exception as e:
        st.error("Processing failed.")
        st.exception(e)

else:
    st.info("Upload both Excel files to begin.")

